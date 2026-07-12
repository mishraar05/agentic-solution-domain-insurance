"""Render the governed source data dictionary as a publishable Excel workbook.

Read-only reporting: consumes dictionary records optionally enriched with
human-reviewed Source Documentation Agent fields and produces a deliverable
file. Never writes to source or recommendation tables.
"""
from collections import OrderedDict

COLUMNS = [
    ("source_column", "Column"),
    ("ordinal_position", "Pos"),
    ("physical_type", "Physical Type"),
    ("nullable", "Nullable"),
    ("key_role", "Key Role"),
    ("proposed_business_name", "Business Name"),
    ("proposed_column_description", "Proposed Column Description"),
    ("proposed_glossary_term", "Proposed Glossary Term"),
    ("proposed_glossary_definition", "Proposed Glossary Definition"),
    ("documentation_generation_status", "Documentation Status"),
    ("documentation_review_state", "Documentation Review"),
    ("ontology_concept_id", "Ontology Concept"),
    ("domain", "Domain"),
    ("observed_or_inferred", "Observed/Inferred"),
    ("privacy_class", "Privacy Class"),
    ("confidence_score", "Confidence"),
    ("approval_state", "Approval State"),
    ("relationship_evidence", "Relationship Evidence"),
    ("assumptions", "Assumptions"),
    ("open_question", "Open Question"),
]


def build_workbook(records, source_system, run_id, artifact_version,
                   approved_only=True):
    """Build an openpyxl Workbook from dictionary record dicts.

    approved_only=True publishes APPROVED records; if False, publishes all
    records with a DRAFT watermark. Returns (workbook, published_count).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    if approved_only:
        records = [r for r in records if r.get("approval_state") == "APPROVED"]
    status = "APPROVED" if approved_only else "DRAFT — CONTAINS UNAPPROVED PROPOSALS"

    tables = OrderedDict()
    for record in sorted(records, key=lambda r: (r["source_table"], r["ordinal_position"])):
        tables.setdefault(record["source_table"], []).append(record)

    wb = Workbook()
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B3139")

    overview = wb.active
    overview.title = "Overview"
    meta_rows = [
        ("Deliverable", "Source Data Dictionary"),
        ("Source system", source_system),
        ("Run ID", run_id),
        ("Artifact version", artifact_version),
        ("Publication status", status),
        ("Tables", len(tables)),
        ("Attributes published", len(records)),
        ("Traceability", "Generated from governed dictionary and source-documentation "
                         "recommendations; every row carries run and review state."),
    ]
    for idx, (key, value) in enumerate(meta_rows, start=1):
        overview.cell(row=idx, column=1, value=key).font = Font(name="Arial", bold=True)
        overview.cell(row=idx, column=2, value=str(value)).font = Font(name="Arial")
    overview.column_dimensions["A"].width = 24
    overview.column_dimensions["B"].width = 90

    for table_name, table_records in tables.items():
        ws = wb.create_sheet(table_name[:31])
        for col_idx, (_, title) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = header_fill
        for row_idx, record in enumerate(table_records, start=2):
            for col_idx, (key, _) in enumerate(COLUMNS, start=1):
                value = record.get(key)
                cell = ws.cell(row=row_idx, column=col_idx,
                               value="" if value is None else value)
                cell.font = Font(name="Arial", size=10)
        ws.freeze_panes = "A2"
        widths = [
            22, 6, 16, 10, 22, 26, 48, 28, 48, 18, 20, 26, 12, 16,
            16, 11, 14, 40, 40, 34,
        ]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    return wb, len(records)
